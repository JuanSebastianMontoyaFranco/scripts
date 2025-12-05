import os
import csv
import time
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from PIL import Image

# --- Configuración ---
TXT_PATH = "Productos_Pendientes.txt"  # archivo con los códigos
INPUT_DIR = "Archivos"  # carpeta donde están tus archivos Excel o CSV
OUTPUT_CSV = "resultados_locales.csv"  # salida
IMG_OUTPUT_DIR = "imagenes"  # carpeta donde se guardarán las imágenes encontradas


def limpia_nombre_archivo(s):
    s = str(s).strip()
    return "".join(c for c in s if c.isalnum() or c in ("-", "_"))


# --- Leer los códigos del TXT ---
with open(TXT_PATH, "r", encoding="utf-8") as f:
    codigos = [line.strip() for line in f if line.strip()]
codigos_set = set(codigos)
print(f"📄 {len(codigos_set)} códigos cargados desde {TXT_PATH}")

os.makedirs(IMG_OUTPUT_DIR, exist_ok=True)


# --- Función para leer archivos Excel o CSV ---
def leer_archivo(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        # Lee todas las hojas del Excel
        xls = pd.ExcelFile(path)
        return {hoja: pd.read_excel(xls, hoja) for hoja in xls.sheet_names}
    elif ext == ".csv":
        return {"Hoja1": pd.read_csv(path, dtype=str)}
    else:
        print(f"⚠️ Formato no compatible: {path}")
        return {}


def extraer_imagenes_coincidencias(ruta_excel, coincidencias_por_hoja):
    """
    Recorre las imágenes incrustadas en el archivo Excel y,
    si la fila donde está la imagen coincide con alguna fila
    donde se encontró un código, guarda la imagen usando ese código.
    """
    try:
        wb = load_workbook(ruta_excel, data_only=True)
    except Exception as e:
        print(f"⚠️ No se pudo abrir '{ruta_excel}' con openpyxl para extraer imágenes: {e}")
        return

    for nombre_hoja, coincidencias_filas in coincidencias_por_hoja.items():
        if nombre_hoja not in wb.sheetnames:
            continue

        ws = wb[nombre_hoja]
        imagenes = getattr(ws, "_images", None)
        if not imagenes:
            continue

        for image in imagenes:
            try:
                celda = image.anchor._from  # openpyxl < 3.1
                fila_img = celda.row + 1    # convertir a base 1
            except AttributeError:
                print(f"⚠️ No pude leer la fila de una imagen en hoja '{nombre_hoja}'. La salto.")
                continue

            # Intentar asociar primero por la misma fila,
            # y si no hay coincidencia, probar con la fila siguiente
            if fila_img in coincidencias_filas:
                codigos_en_fila = coincidencias_filas[fila_img]
            elif (fila_img + 1) in coincidencias_filas:
                codigos_en_fila = coincidencias_filas[fila_img + 1]
            else:
                continue

            try:
                data = image._data()
            except Exception:
                print(f"⚠️ No pude extraer datos binarios de una imagen en hoja '{nombre_hoja}', fila {fila_img}.")
                continue
            # Convertir y guardar siempre como JPG
            try:
                with Image.open(BytesIO(data)) as img:
                    if img.mode in ("RGBA", "P", "LA"):
                        img = img.convert("RGB")

                    for codigo in codigos_en_fila:
                        codigo_str = limpia_nombre_archivo(codigo)
                        nombre_archivo = f"{codigo_str}.jpg"
                        ruta_guardado = os.path.join(IMG_OUTPUT_DIR, nombre_archivo)

                        # Guardar siempre con el mismo nombre; si ya existe, se sobrescribe
                        img.save(ruta_guardado, format="JPEG", quality=95)
                        print(f"📸 Imagen guardada para código {codigo} en: {ruta_guardado}")
            except Exception as e:
                print(f"⚠️ No pude convertir/guardar una imagen en hoja '{nombre_hoja}', fila {fila_img} como JPG: {e}")


# --- Procesar archivos ---
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Archivo", "Hoja", "Fila", "Columna", "Código encontrado"])
    f.flush()

    for nombre_archivo in os.listdir(INPUT_DIR):
        ruta = os.path.join(INPUT_DIR, nombre_archivo)
        if not os.path.isfile(ruta):
            continue

        print(f"🔍 Buscando en: {nombre_archivo}")
        ext = os.path.splitext(ruta)[1].lower()

        # --------------------------
        # Archivos Excel: usar openpyxl
        # --------------------------
        if ext in (".xlsx", ".xlsm", ".xls"):
            try:
                wb = load_workbook(ruta, data_only=True)
            except Exception as e:
                print(f"❌ Error al leer {nombre_archivo} con openpyxl: {e}")
                continue

            # Diccionario: {nombre_hoja: {fila_excel (int): set(códigos_en_esa_fila)}}
            coincidencias_por_hoja = {}

            for nombre_hoja in wb.sheetnames:
                ws = wb[nombre_hoja]
                print(f"   ↳ Revisando hoja: {nombre_hoja}")

                for fila in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        valor = ws.cell(row=fila, column=col).value
                        if valor is None:
                            continue
                        valor_str = str(valor).strip()
                        if not valor_str:
                            continue

                        if valor_str in codigos_set:
                            writer.writerow([nombre_archivo, nombre_hoja, fila, col, valor_str])
                            f.flush()
                            coincidencias_por_hoja.setdefault(nombre_hoja, {}).setdefault(fila, set()).add(valor_str)

            if coincidencias_por_hoja:
                extraer_imagenes_coincidencias(ruta, coincidencias_por_hoja)

        # --------------------------
        # Archivos CSV: usar pandas, sin imágenes
        # --------------------------
        elif ext == ".csv":
            try:
                hojas = leer_archivo(ruta)
            except Exception as e:
                print(f"❌ Error al leer {nombre_archivo}: {e}")
                continue

            for hoja, df in hojas.items():
                print(f"   ↳ Revisando hoja: {hoja}")
                df = df.astype(str)
                filas, columnas = df.shape

                for i in range(filas):
                    for j in range(columnas):
                        valor = str(df.iat[i, j]).strip()
                        if valor in codigos_set:
                            fila_csv = i + 1
                            writer.writerow([nombre_archivo, hoja, fila_csv, j + 1, valor])
                            f.flush()

        else:
            print(f"⚠️ Formato no compatible: {nombre_archivo}")

print(f"✅ Búsqueda completada. Resultados guardados en {OUTPUT_CSV}")
